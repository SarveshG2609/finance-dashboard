"""
Groww XLSX parsers for stock holdings and mutual fund holdings.
- parse_groww_stocks_xlsx: Stocks_Holdings_Statement_*.xlsx  (Sheet1)
- parse_groww_mf_xlsx:     Mutual_Funds_*.xlsx               (Holdings sheet)
"""
import re
from datetime import date
from pathlib import Path

import openpyxl

from app.parsers.base import BrokerHoldingRow, ParsedStatement, SourceKind


# ── Stocks Holdings ──────────────────────────────────────────────────────────

def parse_groww_stocks_xlsx(path: Path, password: str | None = None) -> ParsedStatement:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active or wb.worksheets[0]

    rows_raw = list(ws.iter_rows(values_only=True))
    as_of, invested, closing = _parse_stocks_header(rows_raw)

    holdings: list[BrokerHoldingRow] = []
    header_found = False

    for row in rows_raw:
        if not header_found:
            if _is_stocks_header_row(row):
                header_found = True
            continue
        h = _parse_stocks_row(row, as_of)
        if h:
            holdings.append(h)

    total_value = sum(h.current_or_sell_value for h in holdings)
    total_unrealized = sum(h.unrealized_pnl for h in holdings)

    return ParsedStatement(
        source_kind=SourceKind.BROKER_PNL,
        institution="Groww",
        account_name="Groww Stock Holdings",
        masked_identifier=None,
        statement_start=as_of,
        statement_end=as_of,
        statement_date=as_of,
        summary={
            "invested_value": round(invested or 0, 2),
            "current_value": round(closing or total_value, 2),
            "unrealized_pnl": round(total_unrealized, 2),
            "holdings_count": len(holdings),
        },
        rows=holdings,
        warnings=[],
    )


def _parse_stocks_header(rows: list) -> tuple[date | None, float | None, float | None]:
    as_of = None
    invested = None
    closing = None
    for row in rows:
        cell0 = row[0] if row else None
        cell1 = row[1] if len(row) > 1 else None
        if isinstance(cell0, str) and "as on" in cell0.lower():
            m = re.search(r"(\d{2}-\d{2}-\d{4})", cell0)
            if m:
                try:
                    from datetime import datetime
                    as_of = datetime.strptime(m.group(1), "%d-%m-%Y").date()
                except ValueError:
                    pass
        elif cell0 == "Invested Value" and isinstance(cell1, (int, float)):
            invested = float(cell1)
        elif cell0 == "Closing Value" and isinstance(cell1, (int, float)):
            closing = float(cell1)
    return as_of, invested, closing


def _is_stocks_header_row(row) -> bool:
    return bool(row and row[0] == "Stock Name")


def _parse_stocks_row(row, as_of: date | None) -> BrokerHoldingRow | None:
    # Columns: Stock Name, ISIN, Quantity, Avg buy price, Buy value, Closing price, Closing value, Unrealised P&L
    if not row or not row[0] or not isinstance(row[0], str):
        return None
    name = row[0].strip()
    if name in ("Stock Name", "Total"):
        return None
    isin = str(row[1]).strip() if row[1] else None
    qty = _to_float(row[2]) or 0
    buy_value = _to_float(row[4]) or 0
    current_value = _to_float(row[6]) or 0
    unrealized = _to_float(row[7]) or 0

    if qty <= 0 and current_value <= 0:
        return None

    return BrokerHoldingRow(
        as_of_date=as_of or date.today(),
        symbol_or_name=name,
        isin=isin if isin and isin != "None" else None,
        quantity=qty,
        buy_value=buy_value,
        current_or_sell_value=current_value,
        unrealized_pnl=unrealized,
    )


# ── Mutual Fund Holdings ─────────────────────────────────────────────────────

def parse_groww_mf_xlsx(path: Path, password: str | None = None) -> ParsedStatement:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active or wb.worksheets[0]

    rows_raw = list(ws.iter_rows(values_only=True))
    as_of, total_invested, total_current = _parse_mf_header(rows_raw)

    holdings: list[BrokerHoldingRow] = []
    header_found = False

    for row in rows_raw:
        if not header_found:
            if _is_mf_header_row(row):
                header_found = True
            continue
        h = _parse_mf_row(row, as_of)
        if h:
            holdings.append(h)

    total_value = sum(h.current_or_sell_value for h in holdings)
    total_unrealized = sum(h.unrealized_pnl for h in holdings)

    return ParsedStatement(
        source_kind=SourceKind.BROKER_PNL,
        institution="Groww",
        account_name="Groww Mutual Funds",
        masked_identifier=None,
        statement_start=as_of,
        statement_end=as_of,
        statement_date=as_of,
        summary={
            "invested_value": round(total_invested or 0, 2),
            "current_value": round(total_current or total_value, 2),
            "unrealized_pnl": round(total_unrealized, 2),
            "holdings_count": len(holdings),
        },
        rows=holdings,
        warnings=[],
    )


def _parse_mf_header(rows: list) -> tuple[date | None, float | None, float | None]:
    as_of = None
    total_invested = None
    total_current = None
    for row in rows:
        cell0 = row[0] if row else None
        cell1 = row[1] if len(row) > 1 else None
        if isinstance(cell0, str) and "HOLDINGS AS ON" in cell0.upper():
            m = re.search(r"(\d{4}-\d{2}-\d{2})", cell0)
            if m:
                try:
                    as_of = date.fromisoformat(m.group(1))
                except ValueError:
                    pass
        elif isinstance(cell0, str) and "Total Investments" in cell0:
            # summary row has values in row[1]
            if isinstance(cell1, (int, float)):
                total_invested = float(cell1)
        elif cell0 == "Total Investments" and isinstance(cell1, str):
            # the row after has numeric values at same indices
            pass
    # Also look for the summary values row (str values in cells 0..4)
    for row in rows:
        if row and isinstance(row[0], str):
            try:
                float(str(row[0]).replace(",", ""))
                total_invested = float(str(row[0]).replace(",", ""))
                total_current = float(str(row[1]).replace(",", "")) if row[1] else None
                break
            except (ValueError, TypeError):
                pass
    return as_of, total_invested, total_current


def _is_mf_header_row(row) -> bool:
    return bool(row and row[0] == "Scheme Name")


def _parse_mf_row(row, as_of: date | None) -> BrokerHoldingRow | None:
    # Columns: Scheme Name, AMC, Category, Sub-category, Folio No., Source,
    #          Units, Invested Value, Current Value, Returns, XIRR
    if not row or not row[0] or not isinstance(row[0], str):
        return None
    name = row[0].strip()
    if name in ("Scheme Name",):
        return None

    try:
        invested = float(str(row[7]).replace(",", "")) if row[7] else 0
        current = float(str(row[8]).replace(",", "")) if row[8] else 0
        returns = float(row[9]) if isinstance(row[9], (int, float)) else 0
    except (ValueError, TypeError):
        return None

    if current <= 0 and invested <= 0:
        return None

    return BrokerHoldingRow(
        as_of_date=as_of or date.today(),
        symbol_or_name=name,
        isin=None,
        quantity=float(row[6]) if isinstance(row[6], (int, float)) else 0,
        buy_value=invested,
        current_or_sell_value=current,
        unrealized_pnl=returns,
    )


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None
